import os
import io
from typing import Dict, List, Set, Tuple

import streamlit as st


st.set_page_config(
    page_title="Kindergarten vocabulary",
    page_icon="📚",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
      [data-testid=\"stSidebar\"] { display: none !important; }
      [data-testid=\"collapsedControl\"] { display: none !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Kindergarten vocabulary")


def load_words_from_excel(xlsx_path: str) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    categories: Dict[str, Set[str]] = {
        "Sight Words": set(),
        "Phonetic Words": set(),
        "Family Words": set(),
    }
    family_groups: Dict[str, Set[str]] = {}

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        st.warning("openpyxl is not installed. Run: pip install openpyxl")
        return ({k: [] for k in categories}, {})

    if not os.path.exists(xlsx_path):
        st.error(f"File not found: {xlsx_path}")
        return ({k: [] for k in categories}, {})

    wb = load_workbook(xlsx_path, data_only=True)

    def cat_for(text: str) -> str | None:
        t = (text or "").strip().lower()
        if any(w in t for w in ["sight", "dolch", "fry"]):
            return "Sight Words"
        if any(w in t for w in ["phonic", "phonetic", "phonics"]):
            return "Phonetic Words"
        if any(w in t for w in ["family", "word family", "families"]):
            return "Family Words"
        return None

    def normalize_family_label(label: str) -> str:
        lab = (label or "").strip().lower()
        if not lab:
            return lab
        if lab.startswith("-"):
            return lab
        return f"-{lab}"

    for ws in wb.worksheets:
        # Read header row and normalize
        first_row = next(ws.iter_rows(values_only=True), [])
        headers = [c if isinstance(c, str) else "" for c in first_row]
        headers_lower = [h.strip().lower() for h in headers]

        # Pre-compute indices and sheet category
        word_idx = None
        for i, h in enumerate(headers_lower):
            if h == "word":
                word_idx = i
                break
        sheet_cat = cat_for(ws.title)

        # Detect sheet type via title or header keywords
        sheet_is_sight = (sheet_cat == "Sight Words") or any("sight" in h for h in headers_lower)
        sheet_is_phonetic = (sheet_cat == "Phonetic Words") or any(
            ("phonic" in h) or ("phonetic" in h) or ("phonics" in h) for h in headers_lower
        )

        # A) Exact handling for Sight/Phonetic sheets with a 'word' column: use only that column, skip header
        if sheet_is_sight and word_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if word_idx < len(row):
                    val = row[word_idx]
                    if isinstance(val, str):
                        v = val.strip()
                        if v:
                            categories["Sight Words"].add(v)
            continue

        if sheet_is_phonetic and word_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if word_idx < len(row):
                    val = row[word_idx]
                    if isinstance(val, str):
                        v = val.strip()
                        if v:
                            categories["Phonetic Words"].add(v)
            continue

        # B) Family sheet with both 'family' and 'word' columns: group words by family (skip header)
        fam_idx = None
        for i, h in enumerate(headers_lower):
            if h in ("family", "word family", "rime"):
                fam_idx = i
                break
        if (sheet_cat == "Family Words") and (fam_idx is not None) and (word_idx is not None):
            for row in ws.iter_rows(min_row=2, values_only=True):
                fam = row[fam_idx] if fam_idx < len(row) else None
                w = row[word_idx] if word_idx < len(row) else None
                if isinstance(fam, str) and isinstance(w, str):
                    fam_lab = normalize_family_label(fam)
                    word_val = w.strip()
                    if fam_lab and word_val:
                        family_groups.setdefault(fam_lab, set()).add(word_val)
                        categories["Family Words"].add(word_val)
            continue

        # C) Title-based category: collect all text cells, skipping header row
        if sheet_cat:
            for row in ws.iter_rows(min_row=2, values_only=True):
                for val in row:
                    if isinstance(val, str):
                        v = val.strip()
                        if v:
                            categories[sheet_cat].add(v)
            continue

        # D) Header-based: columns labeled with category keywords. Special-case 'word' for sight/phonetic
        col_cats: Dict[int, str] = {}
        for idx, h in enumerate(headers):
            cat = cat_for(h)
            if cat:
                col_cats[idx] = cat
        if word_idx is not None and sheet_is_sight:
            col_cats[word_idx] = "Sight Words"
        if word_idx is not None and sheet_is_phonetic:
            col_cats[word_idx] = "Phonetic Words"

        if col_cats:
            for row in ws.iter_rows(min_row=2, values_only=True):
                for idx, cat in col_cats.items():
                    if idx < len(row):
                        val = row[idx]
                        if isinstance(val, str):
                            v = val.strip()
                            if v:
                                categories[cat].add(v)

    # If no explicit family groups, infer from common rimes
    if not family_groups and categories["Family Words"]:
        common_rimes = [
            "an","at","ap","am","ad","ag",
            "en","et","ed","eg",
            "in","it","ig","ip","im",
            "on","ot","og","op",
            "un","ut","ug","um",
            "ake","ail","ain","ame","ate","ell","est","ick","ill","ine","ing","ink","ock","ore","uck",
        ]
        for w in categories["Family Words"]:
            wlow = w.lower().strip()
            chosen = None
            for r in sorted(common_rimes, key=len, reverse=True):
                if wlow.endswith(r):
                    chosen = f"-{r}"
                    break
            fam_key = chosen or "-other"
            family_groups.setdefault(fam_key, set()).add(w)

    wb.close()
    categories_out = {k: sorted(list(v)) for k, v in categories.items()}
    family_groups_out = {k: sorted(list(v)) for k, v in family_groups.items()}
    return categories_out, family_groups_out


def render_word_list(words: List[str], columns: int = 3):
    if not words:
        st.info("No entries found.")
        return
    cols = st.columns(columns)
    per_col = (len(words) + columns - 1) // columns
    for i, word in enumerate(words):
        c = cols[(i // per_col) % columns]
        with c:
            st.write(f"- {word}")


# Resolve Excel path relative to project root
excel_path = os.path.abspath(
    os.path.join(os.path.dirname(__file__), os.pardir, "kindergarten_reading_checklist.xlsx")
)
data, family_groups = load_words_from_excel(excel_path)

"""PDF helpers"""
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        ListFlowable,
        ListItem,
        Table,
        TableStyle,
        PageBreak,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

def build_words_pdf_lines(title: str, lines: List[str]) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("PDF engine not available. Install reportlab to enable PDF download.")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54
    )
    title_style = ParagraphStyle(
        name="Title", fontName="Helvetica-Bold", fontSize=16, alignment=TA_LEFT, spaceAfter=12
    )
    item_style = ParagraphStyle(name="Item", fontName="Helvetica", fontSize=12, leading=14)
    story = []
    story.append(Paragraph(title, title_style))
    items = [ListItem(Paragraph(x, item_style), leftIndent=12) for x in lines]
    story.append(ListFlowable(items, bulletType="bullet", start="•"))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def build_words_pdf_columns(title: str, lines: List[str], columns: int = 3) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("PDF engine not available. Install reportlab to enable PDF download.")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54
    )
    title_style = ParagraphStyle(
        name="Title", fontName="Helvetica-Bold", fontSize=16, alignment=TA_LEFT, spaceAfter=12
    )
    cell_style = ParagraphStyle(name="Cell", fontName="Helvetica", fontSize=12, leading=14)

    # Arrange into columns similar to the Streamlit view
    n = len(lines)
    per_col = (n + columns - 1) // columns if n else 0
    cols_data: List[List[str]] = [[] for _ in range(columns)]
    for idx, word in enumerate(lines):
        col_idx = idx // per_col
        if col_idx >= columns:
            col_idx = columns - 1
        cols_data[col_idx].append(word)
    # Pad columns to equal length
    for c in cols_data:
        while len(c) < per_col:
            c.append("")
    # Build row-wise table data
    table_data: List[List[Paragraph]] = []
    for r in range(per_col):
        row: List[Paragraph] = []
        for c in range(columns):
            txt = ("- " + cols_data[c][r]) if cols_data[c][r] else ""
            row.append(Paragraph(txt, cell_style))
        table_data.append(row)

    story = [Paragraph(title, title_style)]
    if table_data:
        tbl = Table(table_data, hAlign="LEFT")
        tbl.setStyle(
            TableStyle(
                [
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# Combined PDF (each section on its own page)
def build_all_pdf(sight: List[str], phon: List[str], family_lines: List[str], grouped: bool) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("PDF engine not available. Install reportlab to enable PDF download.")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54
    )
    title_style = ParagraphStyle(
        name="Title", fontName="Helvetica-Bold", fontSize=16, alignment=TA_LEFT, spaceAfter=12
    )
    cell_style = ParagraphStyle(name="Cell", fontName="Helvetica", fontSize=12, leading=14)
    item_style = ParagraphStyle(name="Item", fontName="Helvetica", fontSize=12, leading=14)

    def section_columns(title: str, lines: List[str], columns: int = 3):
        story = [Paragraph(title, title_style)]
        if not lines:
            return story
        n = len(lines)
        per_col = (n + columns - 1) // columns
        cols_data = [[] for _ in range(columns)]
        for idx, word in enumerate(lines):
            col_idx = min(idx // per_col, columns - 1)
            cols_data[col_idx].append(word)
        for c in cols_data:
            while len(c) < per_col:
                c.append("")
        table_data = []
        for r in range(per_col):
            row = []
            for c in range(columns):
                txt = ("- " + cols_data[c][r]) if cols_data[c][r] else ""
                row.append(Paragraph(txt, cell_style))
            table_data.append(row)
        tbl = Table(table_data, hAlign="LEFT")
        tbl.setStyle(
            TableStyle(
                [
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(tbl)
        return story

    def section_lines(title: str, lines: List[str]):
        story = [Paragraph(title, title_style)]
        if not lines:
            return story
        items = [ListItem(Paragraph(x, item_style), leftIndent=12) for x in lines]
        story.append(ListFlowable(items, bulletType="bullet", start="•"))
        return story

    story: List = []  # type: ignore[var-annotated]
    story += section_columns("Sight Words", sight, columns=3)
    story.append(PageBreak())
    story += section_columns("Phonetic Words", phon, columns=3)
    story.append(PageBreak())
    if grouped:
        story += section_lines("Family Words", family_lines)
    else:
        story += section_columns("Family Words", family_lines, columns=3)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# Prepare data and show combined download at top
_sight = []
_phon = []
_family_lines: List[str] = []
try:
    # data and family_groups are already loaded
    _sight = data.get("Sight Words", [])
    _phon = data.get("Phonetic Words", [])
    _grouped = bool(family_groups)
    if _grouped:
        for fam in sorted(family_groups.keys()):
            words = ", ".join(family_groups[fam])
            _family_lines.append(f"{fam}: {words}")
    else:
        _family_lines = data.get("Family Words", [])
    _any_content = bool(_sight or _phon or _family_lines)
    if _any_content:
        try:
            combined_pdf = build_all_pdf(_sight, _phon, _family_lines, grouped=_grouped)
            st.download_button(
                label="Download All (PDF)",
                data=combined_pdf,
                file_name="kindergarten_all_words.pdf",
                mime="application/pdf",
                key="dl_all_pdf_top",
            )
        except RuntimeError as e:
            st.info(str(e))
            combined_txt = []
            if _sight:
                combined_txt.append("Sight Words:\n" + "\n".join(_sight))
            if _phon:
                combined_txt.append("Phonetic Words:\n" + "\n".join(_phon))
            if _family_lines:
                combined_txt.append("Family Words:\n" + "\n".join(_family_lines))
            st.download_button(
                label="Download All as .txt (fallback)",
                data="\n\n".join(combined_txt),
                file_name="kindergarten_all_words.txt",
                mime="text/plain",
                key="dl_all_txt_top",
            )
except Exception:
    pass
# Sight Words section
st.subheader("Sight Words")
_sight = data.get("Sight Words", [])
render_word_list(_sight)
if _sight:
    try:
        pdf_bytes = build_words_pdf_columns("Sight Words", _sight, columns=3)
        st.download_button(
            label="Download PDF",
            data=pdf_bytes,
            file_name="kindergarten_sight_words.pdf",
            mime="application/pdf",
            key="dl_sight_pdf",
        )
    except RuntimeError as e:
        st.info(str(e))
        st.download_button(
            label="Download as .txt (fallback)",
            data="\n".join(_sight),
            file_name="kindergarten_sight_words.txt",
            mime="text/plain",
            key="dl_sight_txt",
        )

# Phonetic Words section
st.subheader("Phonetic Words")
_phon = data.get("Phonetic Words", [])
render_word_list(_phon)
if _phon:
    try:
        pdf_bytes = build_words_pdf_columns("Phonetic Words", _phon, columns=3)
        st.download_button(
            label="Download PDF",
            data=pdf_bytes,
            file_name="kindergarten_phonetic_words.pdf",
            mime="application/pdf",
            key="dl_phon_pdf",
        )
    except RuntimeError as e:
        st.info(str(e))
        st.download_button(
            label="Download as .txt (fallback)",
            data="\n".join(_phon),
            file_name="kindergarten_phonetic_words.txt",
            mime="text/plain",
            key="dl_phon_txt",
        )

# Family Words section
st.subheader("Family Words")
family_lines: List[str] = []
if family_groups:
    for fam in sorted(family_groups.keys()):
        words = ", ".join(family_groups[fam])
        line = f"{fam}: {words}"
        st.write(line)
        family_lines.append(line)
else:
    _fam = data.get("Family Words", [])
    render_word_list(_fam)
    family_lines = _fam

if family_lines:
    try:
        # Match layout: grouped families are lines; otherwise, 3 columns
        if family_groups:
            pdf_bytes = build_words_pdf_lines("Family Words", family_lines)
        else:
            pdf_bytes = build_words_pdf_columns("Family Words", family_lines, columns=3)
        st.download_button(
            label="Download PDF",
            data=pdf_bytes,
            file_name="kindergarten_family_words.pdf",
            mime="application/pdf",
            key="dl_family_pdf",
        )
    except RuntimeError as e:
        st.info(str(e))
        st.download_button(
            label="Download as .txt (fallback)",
            data="\n".join(family_lines),
            file_name="kindergarten_family_words.txt",
            mime="text/plain",
            key="dl_family_txt",
        )


st.divider()
if st.button("Back to English Vocabulary", type="primary"):
    try:
        st.switch_page("pages/05_English_Vocabulary.py")  # type: ignore[attr-defined]
    except Exception:
        st.warning("Use the Home → English Vocabulary link.")

