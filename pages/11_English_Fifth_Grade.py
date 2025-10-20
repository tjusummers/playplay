import os
from typing import Dict, List, Tuple

import streamlit as st


st.set_page_config(
    page_title="English Vocabulary - Fifth Grade",
    page_icon="V",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# Hide sidebar and its toggle for a clean worksheet-style layout
st.markdown(
    """
    <style>
      [data-testid=\"stSidebar\"] { display: none !important; }
      [data-testid=\"collapsedControl\"] { display: none !important; }
      /* Match Kindergarten button styles */
      [data-testid=\"stDownloadButton\"] > button, .stButton > button {
        background-color: #1f6feb !important;
        color: #ffffff !important;
        border: 1px solid #1f6feb !important;
        border-radius: 6px !important;
        padding: 0.4rem 0.9rem !important;
      }
      [data-testid=\"stDownloadButton\"] > button:hover, .stButton > button:hover { filter: brightness(0.95); }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Fifth Grade Vocabulary")


def load_grade5_vocabulary(xlsx_path: str) -> Tuple[Dict[str, List[str]], List[Tuple[str, str]]]:
    """Load words (and optional definitions) from the grade 5 workbook.

    Returns a tuple:
      - by_sheet: mapping of worksheet name -> list of words
      - definitions: list of (word, definition) tuples (may be empty)

    Heuristics:
      - Prefer a header named 'word' (case-insensitive). If present, use that column.
      - If a 'definition' column exists, pair it with the word.
      - Fallback: collect all string cells from data rows.
    """
    by_sheet: Dict[str, List[str]] = {}
    definitions: List[Tuple[str, str]] = []

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        st.warning("openpyxl is not installed. Run: pip install openpyxl")
        return {}, []

    if not os.path.exists(xlsx_path):
        st.error(f"File not found: {xlsx_path}")
        return {}, []

    wb = load_workbook(xlsx_path, data_only=True)
    try:
        for ws in wb.worksheets:
            # Read header row
            first_row = next(ws.iter_rows(values_only=True), [])
            headers = [c if isinstance(c, str) else "" for c in first_row]
            headers_lower = [h.strip().lower() for h in headers]

            word_idx = None
            def_idx = None
            # Find primary columns
            for i, h in enumerate(headers_lower):
                if h in ("word", "vocabulary", "term") and word_idx is None:
                    word_idx = i
                if h in ("definition", "meaning", "gloss") and def_idx is None:
                    def_idx = i

            words_this_sheet: List[str] = []

            if word_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    w = row[word_idx] if word_idx < len(row) else None
                    if isinstance(w, str):
                        word_val = w.strip()
                        if not word_val:
                            continue
                        words_this_sheet.append(word_val)
                        if def_idx is not None and def_idx < len(row):
                            dv = row[def_idx]
                            if isinstance(dv, str) and dv.strip():
                                definitions.append((word_val, dv.strip()))
            else:
                # Fallback: collect all string cells (skip header row)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for cell in row:
                        if isinstance(cell, str):
                            v = cell.strip()
                            if v:
                                words_this_sheet.append(v)

            # Deduplicate while preserving order
            seen = set()
            deduped: List[str] = []
            for w in words_this_sheet:
                if w not in seen:
                    seen.add(w)
                    deduped.append(w)
            by_sheet[ws.title] = deduped
    finally:
        wb.close()

    return by_sheet, definitions


def render_word_list(words: List[str], columns: int = 3):
    """Display a list of words in N columns, bullet-style (like Kindergarten)."""
    if not words:
        st.info("No entries found.")
        return
    cols = st.columns(columns)
    per_col = (len(words) + columns - 1) // columns
    for i, word in enumerate(words):
        c = cols[(i // per_col) % columns]
        with c:
            st.write(f"- {word}")


# PDF export helpers (used for the top Download All button)
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table
    from reportlab.platypus import TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

def build_words_pdf_columns(title: str, lines: List[str], columns: int = 3) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("PDF engine not available. Install reportlab to enable PDF download.")
    import io
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54
    )
    title_style = ParagraphStyle(
        name="Title", fontName="Helvetica-Bold", fontSize=16, alignment=TA_LEFT, spaceAfter=12
    )
    cell_style = ParagraphStyle(name="Cell", fontName="Helvetica", fontSize=12, leading=14)

    cols = max(1, int(columns))
    per_col = (len(lines) + cols - 1) // cols if lines else 1
    data: List[List[Paragraph]] = []
    for r in range(per_col):
        row_cells: List[Paragraph] = []
        for c in range(cols):
            idx = c * per_col + r
            text = f"- {lines[idx]}" if idx < len(lines) else ""
            row_cells.append(Paragraph(text, cell_style))
        data.append(row_cells)

    table = Table(data)
    table.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ])
    )

    story = [Paragraph(title, title_style), table]
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

def _safe_filename(text: str) -> str:
    s = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in text)
    return s.strip("_") or "export"


# Resolve Excel path relative to project root
excel_path = os.path.abspath(
    os.path.join(os.path.dirname(__file__), os.pardir, "grade5_vocabulary_generic.xlsx")
)
by_sheet, defs = load_grade5_vocabulary(excel_path)

# Aggregate all words
all_words: List[str] = []
for sheet_name in by_sheet:
    all_words.extend(by_sheet[sheet_name])

all_words = list(all_words)

# Top Download All button (PDF), like Kindergarten
if REPORTLAB_AVAILABLE and all_words:
    try:
        pdf_all = build_words_pdf_columns(
            title="Fifth Grade Vocabulary - All Words",
            lines=all_words,
            columns=3,
        )
        # Build quiz PDF (30 random words, 10x3 grid) like Kindergarten
        quiz_pdf = None
        try:
            from reportlab.platypus import Table
            from reportlab.platypus import TableStyle
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_LEFT
            import io, random

            def build_quiz_pdf(words: List[str], rows: int = 10, cols: int = 3) -> bytes:
                from reportlab.platypus import SimpleDocTemplate, Paragraph
                buf = io.BytesIO()
                doc = SimpleDocTemplate(
                    buf, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54
                )
                title_style = ParagraphStyle(
                    name="Title", fontName="Helvetica-Bold", fontSize=16, alignment=TA_LEFT, spaceAfter=12
                )
                label_style = ParagraphStyle(name="Label", fontName="Helvetica", fontSize=12, leading=14)
                cell_style = ParagraphStyle(name="Cell", fontName="Helvetica", fontSize=14, leading=16)

                story: List = []  # type: ignore[var-annotated]
                story.append(Paragraph("Name: ________________________________    Date: ____________", label_style))
                story.append(Paragraph("Fifth Grade Vocabulary Quiz", title_style))

                grid: List[List[Paragraph]] = []
                idx = 0
                for r in range(rows):
                    row: List[Paragraph] = []
                    for c in range(cols):
                        txt = words[idx] if idx < len(words) else ""
                        row.append(Paragraph(txt, cell_style))
                        idx += 1
                    grid.append(row)

                from reportlab.lib.pagesizes import letter as _letter
                available_width = _letter[0] - 54 - 54
                col_w = available_width / cols
                tbl = Table(grid, hAlign="LEFT", colWidths=[col_w] * cols)
                tbl.setStyle(
                    TableStyle(
                        [
                            ("LEFTPADDING", (0, 0), (-1, -1), 6),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("LINEBELOW", (0, 0), (-1, -1), 0.5, "#000000"),
                            ("BOX", (0, 0), (-1, -1), 0.75, "#000000"),
                            ("GRID", (0, 0), (-1, -1), 0.25, "#999999"),
                        ]
                    )
                )
                story.append(tbl)
                doc.build(story)
                buf.seek(0)
                return buf.getvalue()

            pool = list(dict.fromkeys(all_words))
            if pool:
                k = min(30, len(pool))
                picked = random.sample(pool, k=k)
                if k < 30:
                    picked += [""] * (30 - k)
                quiz_pdf = build_quiz_pdf(picked, rows=10, cols=3)
        except Exception:
            quiz_pdf = None

        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                label="Download All (PDF)",
                data=pdf_all,
                file_name=f"grade5_{_safe_filename('all_words')}.pdf",
                mime="application/pdf",
                key="dl_grade5_all_pdf_top",
            )
        with c2:
            if quiz_pdf is not None:
                st.download_button(
                    label="Create a Quiz",
                    data=quiz_pdf,
                    file_name="grade5_vocab_quiz.pdf",
                    mime="application/pdf",
                    key="dl_grade5_quiz_pdf_top",
                )
    except Exception:
        st.info("PDF export unavailable. Install reportlab if needed.")

st.subheader("All Words")
render_word_list(all_words, columns=3)

# Per-sheet sections (mirrors category blocks in Kindergarten)
for sheet_name, words in by_sheet.items():
    shown = list(words)
    if not shown:
        continue
    st.subheader(sheet_name)
    render_word_list(shown, columns=3)

# Optional definitions table if present
if defs:
    st.subheader("Definitions")
    rows = list(defs)
    st.dataframe({"Word": [w for w, _ in rows], "Definition": [d for _, d in rows]}, use_container_width=True)


st.divider()
if st.button("Back to English Vocabulary", type="primary"):
    try:
        st.switch_page("pages/05_English_Vocabulary.py")  # type: ignore[attr-defined]
    except Exception:
        st.warning("Open the English Vocabulary page from Home.")
